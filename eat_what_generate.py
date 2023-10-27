from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import ttk
from pandastable import Table
from random import randint

def display_excel_data(file_path):
    # Read the Excel file using pandas
    df = pd.read_excel(file_path)

    # Create a new window
    window = tk.Tk()
    window.title("Eat What Generator")

    # Determine the number of stores
    num_stores = len(df)

    # Generate a random number
    random_number = randint(1, num_stores)

    # Get the store name at the randomly generated index
    store_name = df.iloc[random_number - 1]['store']

    # Create a label frame for store information
    info_frame = ttk.LabelFrame(window, text="Store Information")
    info_frame.pack(pady=10)

    # Create labels to display store information
    no_store_label = tk.Label(info_frame, text=f"Number of Stores: {num_stores}")
    no_store_label.pack()

    random_label = tk.Label(info_frame, text=f"Random Number: {random_number}")
    random_label.pack(pady=5)

    store_label = tk.Label(info_frame, text=f"Store Name: {store_name}")
    store_label.pack()

    # Create a frame for the table
    table_frame = ttk.Frame(window)
    table_frame.pack(fill='both', expand=True)

    # Create a PandasTable widget
    table = Table(table_frame, dataframe=df)
    table.show()

    # Update the Excel file by incrementing the cell value related to the randomly generated store
    increment_cell_value(file_path, random_number, 2)

    fit_columns_to_header(table)

    # Refresh the table after updating the Excel file
    table.model.df = pd.read_excel(file_path)
    table.redraw()


    def redo_operation():
        nonlocal random_number, store_name

        # Generate a new random number within the range of the number of stores
        random_number = randint(1, num_stores)

        # Get the store name at the newly generated index
        store_name = df.iloc[random_number - 1]['store']

        # Update the labels
        random_label.config(text=f"Random Number: {random_number}")
        store_label.config(text=f"Store Name: {store_name}")

        # Update the Excel file by incrementing the cell value related to the randomly generated store
        increment_cell_value(file_path, random_number, 2)

        # Refresh the table after updating the Excel file
        table.model.df = pd.read_excel(file_path)
        table.redraw()

    def select_operation():
        # Update the Excel file by incrementing the cell value related to the randomly generated store
        increment_cell_value(file_path, random_number, 3)

        # Refresh the table after updating the Excel file
        table.model.df = pd.read_excel(file_path)
        table.redraw()


    # Create a button frame for redo and select buttons
    button_frame = ttk.Frame(window)
    button_frame.pack(pady=10)

    # Create a redo button
    redo_button = ttk.Button(button_frame, text="Redo", command=redo_operation)
    redo_button.pack(side='left', padx=5)

    # Create a select next button
    select_button = ttk.Button(button_frame, text="Select", command=select_operation)
    select_button.pack(side='left', padx=5)

    # Run the window's main loop
    window.mainloop()

def increment_cell_value(file_path, random_number, column):
    # Load the Excel file using openpyxl
    wb = load_workbook(file_path)

    # Select the appropriate worksheet (replace 'Sheet1' with the actual sheet name)
    ws = wb['Sheet1']

    # Increment the cell value related to the randomly generated store
    cell = ws.cell(row=random_number + 1, column=column)  # Assuming the store numbers are in column B (column index 2)
    cell.value += 1

    # Save the updated Excel file
    wb.save(file_path)

def fit_columns_to_header(table):
    # Get the header labels
    header_labels = table.model.df.columns.tolist()

    # Get the maximum width for each column based on the header content
    max_widths = [len(str(label)) for label in header_labels]

    # Get the cell content for each column and update the maximum width if necessary
    for i, column in enumerate(table.model.df.columns):
        column_values = table.model.df[column].tolist()
        max_widths[i] = max(max_widths[i], max([len(str(value)) for value in column_values]))

    # Set the column widths to match the maximum width for each column
    for i, width in enumerate(max_widths):
        table.columnconfigure(i, minsize=width * 10)  # Adjust the width factor as needed


# Example usage
file_path = '..\\eat what generater\\store.xlsx'
display_excel_data(file_path)